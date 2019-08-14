from table_handler import first_blank_row_finder
from openpyxl import load_workbook # Работа с таблицами

from save_numbers import read_numbers
number_of_people, number_of_vegans = read_numbers()  #Количество людей и вегетарианцев среди них

def ingredients_calculator(to_cook):
    # ВХОД: список блюд (словарь to_cook)
    #        количество людей и вегетарианцев среди них (number_of_people, number_of_vegans),
    #        таблица с количествами каждого ингредиента в блюде на человека (ingredients.xslx),
    # ВЫХОД: список продуктов для покупки (словарь to_buy),
    #        список ненайденных в таблице ingredients блюд (unknown_dishes)

    to_buy = dict()
    unknown_dishes = set()

    # Импортируем таблицу с ингредиентами
    ingredients = load_workbook(filename='ingredients.xlsx')['Ingredients']
    # Находим перву пустую строку
    ingredients_first_blank_row = first_blank_row_finder(ingredients, 3)

    # В dish_names запишем все названия блюд (из первого столбца)
    dish_names = [ingredients.cell(row=i, column=1) for i in range(1, ingredients_first_blank_row)]

    # Функция ниже находит нужное название блюда в списке блюд и выдаёт его строку
    # Если блюдо не найдено, выдаёт not_found
    def search_dish(name):
        for cell in dish_names:
            if cell.value == name:
                return cell.row
        return 'not_found'

    # Функция ниже получает на вход строку блюда.
    # Она показывает, сколько строк ингредиентов относится к этому блюду
    def search_next_dish(row):
        end = 1
        value = None
        while (value == None) and (row + end <= ingredients_first_blank_row):
            value = ingredients.cell(column=1, row=row+end).value
            end += 1
        return end

    # Обрабатываем блюда в списке to_cook
    for dish in to_cook:
        row = search_dish(dish)
        # Находим ряд блюда
        if row == 'not_found':
            # Если блюда нет в таблице ingredients
            # То мы добавим его в список ненайденных блюд
            unknown_dishes.add(dish)
        else:
            # Если же оно там есть, то
            # умножим каждый ингредиент на число людей и количество блюда
            for i in range(row, row + search_next_dish(row) - 1):
                ingredient = ingredients.cell(column=2,row=i).value
                quantity = ingredients.cell(column=3,row=i).value * to_cook[dish] * number_of_people
                if ingredient in to_buy:
                    # Если этот ингредиент уже есть в словаре
                    # Мы сложим количество этого ингредиента с уже имеющимся
                    to_buy[ingredient] += quantity
                else:
                    # Если нет, то добавим этот ингредиент в словарь
                    to_buy[ingredient] = quantity

    return to_buy, unknown_dishes,